import pandas as pd
import holidays
from meteostat import Point, Hourly
from openpyxl.utils import get_column_letter

def getWeatherData(latitude: float, longitude: float, start_date: pd.Timestamp, end_date: pd.Timestamp) -> pd.DataFrame:

    location = Point(latitude, longitude)
    data = Hourly(location, start_date, end_date).fetch()
    
    weather_df = data[['temp', 'rhum']].rename(columns={'temp': 'Temperatura Externa', 'rhum': 'Umidade'})
    weather_df = weather_df.resample('10min').interpolate(method='linear')
    
    return weather_df

def applySheetsFormatting(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):

    worksheet = writer.sheets[sheet_name]
    
    for idx, col_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(idx)
        if "Timestamp" in col_name:
            max_len = 19
        else:
            max_len = max(df[col_name].astype(str).map(len).max(), len(col_name))
        worksheet.column_dimensions[column_letter].width = max_len + 2

    formato_kwh = '0.00'
    formato_int = '0'
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        if 'Consumo' in col_name:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = formato_kwh
        elif col_name in ['Temperatura Externa', 'Umidade'] or col_name.startswith('boolean'):
            for cell in worksheet[col_letter][1:]:
                cell.number_format = formato_int

def parametersPipeline():

    FILE_ENTRADA = 'consumoEnergiaEletrica.xlsx'
    FILE_SAIDA = 'consumoEnergiaEletrica_Parametros.xlsx'
    LATITUDE, LONGITUDE = -X, Y

    print(f"Iniciando implementação de parâmetros de otimização do arquivo '{FILE_ENTRADA}'.")
    
    try:
        abas_originais = pd.read_excel(FILE_ENTRADA, sheet_name=None)
        print(f"Arquivo lido com sucesso. {len(abas_originais)} meses detectados.")
    except FileNotFoundError:
        print(f"ERRO: O arquivo de entrada '{FILE_ENTRADA}' não foi encontrado.")
        return

    allYears = set()
    for df in abas_originais.values():
        if 'Timestamp' in df.columns:
            allYears.update(pd.to_datetime(df['Timestamp']).dt.year.unique())
    feriadosEstaduais = holidays.BR(state='RJ', years=list(allYears))
    

    with pd.ExcelWriter(FILE_SAIDA, engine='openpyxl') as writer:
        for sheet_name, df in abas_originais.items():
            print(f"Atualizando com novos parâmetros: {sheet_name}...")
            
            df['Timestamp'] = pd.to_datetime(df['Timestamp'])
            
            start_date, end_date = df['Timestamp'].min(), df['Timestamp'].max()
            weather_df = getWeatherData(LATITUDE, LONGITUDE, start_date, end_date)
            
            df_parametros = pd.merge_asof(df.sort_values('Timestamp'), weather_df, left_on='Timestamp', right_index=True)
            df_parametros ['Hora'] = df_parametros['Timestamp'].dt.hour

            df_parametros['booleanDiaUtil'] = (df_parametros['Timestamp'].dt.weekday < 5).astype(int)
            df_parametros['booleanFeriado'] = df_parametros['Timestamp'].dt.date.isin(feriadosEstaduais).astype(int)
            df_parametros.loc[df_parametros['booleanFeriado'] == 1, 'boleanDiaUtil'] = 0
            
            df_parametros['booleanPicoPredio'] = df_parametros['Hora'].isin([10, 11]).astype(int)
            df_parametros['booleanPicoConcessionaria'] = df_parametros['Hora'].between(17, 20).astype(int)
            
            df_parametros['Temperatura Externa'] = df_parametros['Temperatura Externa'].round(0).astype('Int64')
            df_parametros['Umidade'] = df_parametros['Umidade'].round(0).astype('Int64')
            
            consumo_cols = [c for c in df.columns if 'Consumo' in c]
            final_cols = [
                'Timestamp', 'Temperatura Externa', 'Umidade',
                'booleanDiaUtil', 'booleanFeriado', 'booleanPicoPredio', 'booleanPicoConcessionaria',
            ] + consumo_cols
            
            df_final = df_parametros[[col for col in final_cols if col in df_parametros.columns]]
            
            df_final.to_excel(writer, sheet_name=sheet_name, index=False)
            applySheetsFormatting(writer, sheet_name, df_final)
            
    print(f"\nProcedimento concluído. Arquivo final '{FILE_SAIDA}' foi gerado com sucesso.")

if __name__ == "__main__":
    parametersPipeline()